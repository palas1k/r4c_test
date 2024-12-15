import datetime
from io import BytesIO
from pprint import pprint

from django.db.models import Count
from openpyxl import Workbook

from robots.models import Robot


class DownloadRobots:
    end_date = datetime.datetime.now()
    start_date = end_date - datetime.timedelta(days=7)

    def _get_robots_info(self):
        robots = list(Robot.objects.filter(created__range=(self.start_date, self.end_date)).order_by('model'))
        return robots

    def render_excel_file(self):
        robots_info = self._get_robots_info()
        workbook = Workbook()
        sheets = list()
        header = ["Модель", "Версия", "Количество за неделю"]
        for robot in robots_info:
            if robot.model not in sheets:
                sheet = workbook.create_sheet(title=robot.model)
                sheet.append(header)
                sheets.append(robot.model)
            sheet.append([robot.model, robot.version, robot.quantity])
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        output.seek(0)
        return output
